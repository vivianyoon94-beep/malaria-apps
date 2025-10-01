
import io
import pandas as pd
import streamlit as st

# Import library functions
from Malaria_Data_Cleaning import clean_malaria_data
from Malaria_Indicator import compute_indicators

# --- new: openpyxl for preserving untouched sheets ---
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- Utilities ----------------
def put_comment_first(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[["COMMENT"] + [c for c in df.columns if c != "COMMENT"]]

def put_comment_last(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[[c for c in df.columns if c != "COMMENT"] + ["COMMENT"]]

def sanitize_sheet_name(name: str) -> str:
    bad = '[]:*?/\\'
    table = {ord(ch): ' ' for ch in bad}
    name = str(name).translate(table).strip() or "Sheet"
    return name[:31]

# ---------- Chip-free multi-select dropdown ----------
def multiselect_dropdown(label: str, options: list[str], key: str):
    sel_key  = f"{key}_selected"
    snap_key = f"{key}_options_snapshot"

    if sel_key not in st.session_state:
        st.session_state[sel_key] = set()
    if snap_key not in st.session_state:
        st.session_state[snap_key] = tuple()

    snapshot = tuple(options)
    if st.session_state[snap_key] != snapshot:
        st.session_state[snap_key] = snapshot
        st.session_state[sel_key] = set()

    def render_checkbox_list():
        cols = st.columns([1, 1])
        with cols[0]:
            if st.button("Select all", key=f"{key}_all"):
                st.session_state[sel_key] = set(options)
        with cols[1]:
            if st.button("Clear", key=f"{key}_clear"):
                st.session_state[sel_key].clear()

        for opt in options:
            checked = opt in st.session_state[sel_key]
            if st.checkbox(opt, value=checked, key=f"{key}_{opt}"):
                st.session_state[sel_key].add(opt)
            else:
                st.session_state[sel_key].discard(opt)

    if hasattr(st, "popover"):
        with st.popover(label):
            render_checkbox_list()
    else:
        with st.expander(label, expanded=False):
            render_checkbox_list()

    selected = [o for o in options if o in st.session_state[sel_key]]
    st.caption(f"Selected: {len(selected)}")
    return selected

# Optional helper: normalize SCREENING_DATE as 'DD-MMM-YY' for display
import re as _re
from datetime import datetime as _dt
_ISO_DATE = _re.compile(r'^\d{4}[-/]\d{2}[-/]\d{2}(?: \d{2}:\d{2}:\d{2})?$')
def force_screening_date_strings(df: pd.DataFrame) -> pd.DataFrame:
    date_col = next((c for c in df.columns if c.strip().lower() == "screening_date"), None)
    if not date_col:
        return df
    def _fmt(v):
        if isinstance(v, (pd.Timestamp, _dt)):
            return pd.to_datetime(v).strftime('%d-%b-%y')
        if isinstance(v, str) and _ISO_DATE.match(v.strip()):
            dt = pd.to_datetime(v, errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%d-%b-%y')
        return v
    df[date_col] = df[date_col].map(_fmt)
    return df

# ----------------- App -----------------
st.set_page_config(page_title="🦟 Malaria Apps", layout="wide")
st.title("🦟 Malaria Apps")


# --- Section 0: Sheet Merger ---
st.header("Sheet Merger")

merge_file = st.file_uploader("Upload an Excel file to merge sheets", type=["xlsx", "xls"], key="merge_uploader")

def _normalize_headers(df: pd.DataFrame):
    return [str(c).strip().lower() for c in df.columns]

def _validate_headers_match(dfs):
    base = set(_normalize_headers(dfs[0]))
    for i, d in enumerate(dfs[1:], start=2):
        if set(_normalize_headers(d)) != base:
            return False, i
    return True, None

if merge_file is not None:
    try:
        merge_xls = pd.ExcelFile(merge_file)
        merge_sheets_all = merge_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file. Please check the format.")
        st.exception(e)
        merge_xls = None
        merge_sheets_all = []

    if merge_sheets_all:
        merge_selected = multiselect_dropdown("Select sheet(s) to merge", merge_sheets_all, key="merge")
        st.write("")
        include_originals = st.checkbox("Include original sheets in output (preserve formatting where possible)", value=True)
        run_merge = st.button("Run", key="run_merge")

        if run_merge:
            if not merge_selected or len(merge_selected) < 2:
                st.warning("Select at least **two** sheets to merge, then click **Run**.")
            else:
                # Read selected sheets
                sheet_dfs = []
                for s in merge_selected:
                    df = merge_xls.parse(sheet_name=s)
                    df.columns = [str(c) for c in df.columns]  # ensure string col names
                    sheet_dfs.append((s, df))

                ok, bad_idx = _validate_headers_match([d for _, d in sheet_dfs])
                if not ok:
                    bad_name = merge_selected[bad_idx-1] if bad_idx and bad_idx-1 < len(merge_selected) else "Unknown"
                    st.error(f"Headers do not match across selected sheets. Mismatch around sheet: **{bad_name}**")
                else:
                    # Canonical order based on first sheet
                    canonical = list(sheet_dfs[0][1].columns)
                    merged_parts = []
                    for s, df in sheet_dfs:
                        # reorder columns case-insensitively to match canonical
                        mapping = {str(c).strip().lower(): c for c in df.columns}
                        ordered = [mapping[str(c).strip().lower()] for c in canonical]
                        tmp = df[ordered].copy()
                        tmp["DATA_SOURCE"] = s
                        merged_parts.append(tmp)
                    merged_df = pd.concat(merged_parts, ignore_index=True)

                    st.subheader("👀 Merged preview (first 50 rows)")
                    st.dataframe(merged_df.head(50), use_container_width=True)

                    # Download
                    st.subheader("⬇️ Download")
                    try:
                        if include_originals and str(merge_file.name).lower().endswith(".xlsx"):
                            # Use openpyxl to append merged sheet to original, preserving untouched formatting
                            original_bytes = merge_file.getvalue() if hasattr(merge_file, "getvalue") else merge_file.read()
                            wb = load_workbook(io.BytesIO(original_bytes))

                            # ensure a unique sheet name
                            base_name = "Merged"
                            name = base_name
                            counter = 1
                            while name in wb.sheetnames:
                                counter += 1
                                name = f"{base_name}_{counter}"
                            ws = wb.create_sheet(title=name)
                            for r in dataframe_to_rows(merged_df, index=False, header=True):
                                ws.append(r)
                            out = io.BytesIO()
                            wb.save(out); out.seek(0)
                            st.download_button(
                                label=f"📥 Download with original sheets + merged ('{name}')",
                                data=out.getvalue(),
                                file_name="merged_with_originals.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                        else:
                            # Only merged sheet in a fresh workbook
                            out = io.BytesIO()
                            with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
                                merged_df.to_excel(writer, index=False, sheet_name="Merged")
                            out.seek(0)
                            st.download_button(
                                label="📥 Download merged-only workbook",
                                data=out.getvalue(),
                                file_name="merged_only.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                    except Exception as e:
                        st.error("❌ Failed to build the merged workbook.")
                        st.exception(e)

st.markdown("---")
# --- Section 1: Data Cleaning ---
st.header("Data Cleaning")

clean_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="clean_uploader")

if clean_file is not None:
    try:
        clean_xls = pd.ExcelFile(clean_file)
        clean_sheets = clean_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file. Please check the format.")
        st.exception(e)
        clean_xls = None
        clean_sheets = []

    if clean_sheets:
        clean_selected = multiselect_dropdown("Select sheet(s) to process", clean_sheets, key="clean")
        st.write("")  # small gap
        run_clean = st.button("Run", key="run_clean")

        if run_clean:
            if not clean_selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                # Process selected sheets to generate cleaned DataFrames
                cleaned_by_sheet = {}
                for sheet in clean_selected:
                    raw_df  = clean_xls.parse(sheet_name=sheet)
                    cleaned = clean_malaria_data(raw_df)
                    cleaned = force_screening_date_strings(cleaned)
                    cleaned_by_sheet[sheet] = {
                        "display": put_comment_first(cleaned.copy()),
                        "file":    put_comment_last(cleaned.copy()),
                    }

                st.subheader("⚠️ Error Rows Preview (by sheet)")
                tabs = st.tabs(clean_selected)

                def _build_error_summary_by_column(error_series: pd.Series) -> pd.DataFrame:
                    counts = {}
                    for comment in error_series.astype(str):
                        parts = [p.strip() for p in comment.split(';') if p.strip()]
                        for p in parts:
                            if '[' in p and ']' in p:
                                col = p.split('[', 1)[1].split(']', 1)[0]
                                counts[col] = counts.get(col, 0) + 1
                    if not counts:
                        return pd.DataFrame(columns=['Column', 'Error Count'])
                    return (
                        pd.DataFrame(list(counts.items()), columns=['Column', 'Error Count'])
                        .sort_values('Error Count', ascending=False)
                        .reset_index(drop=True)
                    )

                for tab, sheet in zip(tabs, clean_selected):
                    with tab:
                        display_df = cleaned_by_sheet[sheet]["display"]
                        error_df = display_df[display_df.get("COMMENT", "").astype(str).str.strip() != ""]
                        if error_df.empty:
                            st.success(f"✅ {sheet}: No data quality issues found.")
                        else:
                            st.dataframe(error_df.head(50), use_container_width=True)
                            st.info(f"{sheet} — Total error rows: {len(error_df)}")
                            summary_df = _build_error_summary_by_column(error_df["COMMENT"])
                            if not summary_df.empty:
                                st.markdown("**📊 Error Summary by Column**")
                                st.dataframe(summary_df, use_container_width=True)

                # --- Build downloadable workbook, preserving UNTOUCHED sheets via openpyxl ---
                st.subheader("⬇️ Download")
                try:
                    # Read original bytes and load workbook so untouched sheets keep styles/filters
                    original_bytes = clean_file.getvalue() if hasattr(clean_file, "getvalue") else clean_file.read()
                    wb = load_workbook(io.BytesIO(original_bytes))

                    # For each selected/cleaned sheet: replace content (formatting will be new for these sheets)
                    for sheet in clean_xls.sheet_names:
                        if sheet in cleaned_by_sheet:
                            # Remove existing sheet, insert a new plain sheet at the same position
                            try:
                                idx = wb.sheetnames.index(sheet)
                            except ValueError:
                                idx = len(wb.sheetnames)
                            ws_old = wb[sheet] if sheet in wb.sheetnames else None
                            if ws_old is not None:
                                wb.remove(ws_old)
                            ws = wb.create_sheet(title=sheet, index=idx)

                            # Write DataFrame (header + rows)
                            for r in dataframe_to_rows(cleaned_by_sheet[sheet]["file"], index=False, header=True):
                                ws.append(r)
                        # else: untouched sheet remains exactly as-is

                    # Save to buffer
                    out = io.BytesIO()
                    wb.save(out)
                    out.seek(0)

                    st.download_button(
                        label=f"📥 Download Cleaned Workbook",
                        data=out.getvalue(),
                        file_name="malaria_cleaned_all.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error("❌ Failed to generate the cleaned workbook.")
                    st.exception(e)

st.markdown("---")

# --- Section 2: Indicator Calculation ---
st.header("Indicator Calculation")

ind_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="ind_uploader")

if ind_file is not None:
    try:
        ind_xls = pd.ExcelFile(ind_file)
        ind_sheets = ind_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file. Please check the format.")
        st.exception(e)
        ind_xls = None
        ind_sheets = []

    if ind_sheets:
        ind_selected = multiselect_dropdown("Select sheet(s) to process", ind_sheets, key="ind")
        st.write("")  # small gap
        run_ind = st.button("Run", key="run_ind")

        if run_ind:
            if not ind_selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                outputs = {}
                errors = []
                for sheet in ind_selected:
                    try:
                        raw_df = ind_xls.parse(sheet_name=sheet)
                        out_df = compute_indicators(raw_df.copy())
                        outputs[sheet] = out_df
                    except Exception as e:
                        errors.append((sheet, e))

                if outputs:
                    st.subheader("👀 Preview (first 50 rows)")
                    tabs = st.tabs(list(outputs.keys()))
                    for tab, sheet in zip(tabs, outputs.keys()):
                        with tab:
                            st.caption(f"Sheet: {sheet}")
                            st.dataframe(outputs[sheet].head(50), use_container_width=True)

                if errors:
                    st.warning("Some sheets failed to process:")
                    for sheet, e in errors:
                        with st.expander(f"Details: {sheet}"):
                            st.exception(e)

                # --- Build downloadable workbook, preserving UNTOUCHED sheets via openpyxl ---
                if outputs or ind_selected:
                    st.subheader("⬇️ Download")
                    try:
                        original_bytes = ind_file.getvalue() if hasattr(ind_file, "getvalue") else ind_file.read()
                        wb = load_workbook(io.BytesIO(original_bytes))

                        for sheet in ind_xls.sheet_names:
                            if sheet in outputs:
                                try:
                                    idx = wb.sheetnames.index(sheet)
                                except ValueError:
                                    idx = len(wb.sheetnames)
                                ws_old = wb[sheet] if sheet in wb.sheetnames else None
                                if ws_old is not None:
                                    wb.remove(ws_old)
                                ws = wb.create_sheet(title=sheet, index=idx)

                                for r in dataframe_to_rows(outputs[sheet], index=False, header=True):
                                    ws.append(r)
                            # else: untouched sheet remains exactly as-is

                        out = io.BytesIO()
                        wb.save(out)
                        out.seek(0)

                        st.download_button(
                            label=f"📥 Download Indicators Workbook",
                            data=out.getvalue(),
                            file_name="malaria_indicators_all.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error("❌ Failed to generate the indicators workbook.")
                        st.exception(e)
