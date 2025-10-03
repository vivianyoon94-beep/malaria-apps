import io
import zipfile
import pandas as pd
import streamlit as st

# your cleaning/indicator modules
from Malaria_Data_Cleaning import clean_malaria_data
from Malaria_Indicator import compute_indicators

# preserve untouched formatting for .xlsx writes / injection
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------- small helpers ----------
def put_comment_first(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[["COMMENT"] + [c for c in df.columns if c != "COMMENT"]]

def put_comment_last(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[[c for c in df.columns if c != "COMMENT"] + ["COMMENT"]]

def _normalize_headers(df: pd.DataFrame):
    return [str(c).strip().lower() for c in df.columns]

def _validate_headers_match(dfs):
    if not dfs:
        return False, "No dataframes selected"
    base = set(_normalize_headers(dfs[0]))
    for i, d in enumerate(dfs[1:], start=2):
        if set(_normalize_headers(d)) != base:
            return False, i
    return True, None

def _strip_time_from_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """For true datetime64 columns, keep only the date value (for writing)."""
    for col, dtype in df.dtypes.items():
        if pd.api.types.is_datetime64_any_dtype(dtype):
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
    return df

def _format_worksheet_dates(ws, df: pd.DataFrame, header_row: int = 1, first_data_row: int = 2):
    """
    On an openpyxl worksheet that already has the DataFrame written,
    set Excel number format 'DD-MMM-YY' for columns that contain real date/datetime values.
    """
    from datetime import date as _dt_date
    date_cols = []
    for j, col in enumerate(df.columns, start=1):
        s = df[col]
        if pd.api.types.is_datetime64_any_dtype(s) or s.map(lambda v: isinstance(v, (_dt_date, pd.Timestamp))).any():
            date_cols.append(j)
    for j in date_cols:
        for col_cells in ws.iter_cols(min_col=j, max_col=j, min_row=first_data_row):
            for cell in col_cells:
                cell.number_format = "DD-MMM-YY"

def _df_display_without_time(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a copy for Streamlit preview where datetime64 columns are rendered as 'DD-MMM-YY'.
    Does not touch string columns (e.g., '08.4.25' stays as typed).
    """
    out = df.copy()
    for col, dtype in out.dtypes.items():
        if pd.api.types.is_datetime64_any_dtype(dtype):
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%d-%b-%y")
    return out


# ----------------- app -----------------
st.set_page_config(page_title="🦟 Malaria App", layout="wide")
st.title("🦟 Malaria App")

# === Section 0: Sheet Merger (single or multiple files) ===
st.header("Sheet Merger (single or multiple files)")

merge_files = st.file_uploader(
    "Upload one or more Excel files to merge",
    type=["xlsx", "xls"],
    key="merge_uploader_multi",
    accept_multiple_files=True,
)

if merge_files:
    file_objs = []
    for f in merge_files:
        try:
            xls = pd.ExcelFile(f)
            file_objs.append((f, xls))
        except Exception as e:
            st.error(f"❌ Could not read: {f.name}")
            st.exception(e)
            file_objs.append((f, None))

    selections = {}
    for f, xls in file_objs:
        st.subheader(f"File: {f.name}")
        if xls is None:
            continue
        sheets = xls.sheet_names
        chosen = st.multiselect(f"Select sheet(s) from **{f.name}** to include", sheets, key=f"sel_{f.name}")
        selections[f.name] = {"sheets": chosen, "xls": xls, "file": f}

    st.write("")
    run_merge = st.button("Run", key="run_merge_multi")

    if run_merge:
        parts = []
        for fname, meta in selections.items():
            xls = meta.get("xls")
            chosen = meta.get("sheets", [])
            if xls is None or not chosen:
                continue
            for s in chosen:
                df = xls.parse(sheet_name=s)
                df.columns = [str(c) for c in df.columns]
                df["DATA_SOURCE"] = s
                df["FILE_SOURCE"] = fname
                parts.append(df)

        if not parts:
            st.warning("Please select at least one sheet to merge.")
        else:
            ok, info = _validate_headers_match(parts)
            if not ok:
                msg = f"Headers do not match (mismatch near input #{info})." if isinstance(info, int) else str(info)
                st.error(msg)
            else:
                meta_cols = {"DATA_SOURCE", "FILE_SOURCE"}
                canonical = [c for c in parts[0].columns if c not in meta_cols]
                normalized = []
                for df in parts:
                    mapping = {str(c).strip().lower(): c for c in df.columns}
                    ordered = [mapping[str(c).strip().lower()] for c in canonical]
                    tmp = df[ordered + ["DATA_SOURCE", "FILE_SOURCE"]].copy()
                    normalized.append(tmp)
                merged_df = pd.concat(normalized, ignore_index=True)
                merged_df = _strip_time_from_datetime_columns(merged_df)

                st.subheader("👀 Merged preview (first 50 rows)")
                st.dataframe(_df_display_without_time(merged_df).head(50), use_container_width=True)

                # ---- Downloads ----
                st.subheader("⬇️ Download")

                # 1) merged-only workbook (date format set)
                out = io.BytesIO()
                with pd.ExcelWriter(
                    out,
                    engine="xlsxwriter",
                    datetime_format="dd-mmm-yy",
                    date_format="dd-mmm-yy",
                ) as writer:
                    merged_df.to_excel(writer, index=False, sheet_name="Merged")
                out.seek(0)
                st.download_button(
                    label="📥 Download merged-only workbook",
                    data=out.getvalue(),
                    file_name="merged_across_files.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                # 2) ZIP with original files + merged sheet (first, date format columns)
                zipbuf = io.BytesIO()
                with zipfile.ZipFile(zipbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    from datetime import date as _dt_date
                    for f, xls in file_objs:
                        fname = f.name
                        if fname.lower().endswith(".xlsx"):
                            orig_bytes = f.getvalue() if hasattr(f, "getvalue") else f.read()
                            wb = load_workbook(io.BytesIO(orig_bytes))

                            base_name = "Merged"
                            name = base_name
                            counter = 1
                            while name in wb.sheetnames:
                                counter += 1
                                name = f"{base_name}_{counter}"
                            ws = wb.create_sheet(title=name, index=0)

                            for r in dataframe_to_rows(merged_df, index=False, header=True):
                                ws.append(r)

                            # Format date columns for display
                            date_cols_idx = []
                            for j, col in enumerate(merged_df.columns, start=1):
                                col_series = merged_df[col]
                                if pd.api.types.is_datetime64_any_dtype(col_series) or col_series.map(
                                    lambda v: isinstance(v, (_dt_date, pd.Timestamp))
                                ).any():
                                    date_cols_idx.append(j)
                            for j in date_cols_idx:
                                for col_cells in ws.iter_cols(min_col=j, max_col=j, min_row=2):
                                    for cell in col_cells:
                                        cell.number_format = "DD-MMM-YY"

                            fout = io.BytesIO()
                            wb.save(fout); fout.seek(0)
                            zf.writestr(fname, fout.read())
                        else:
                            # .xls fallback: new xlsx with merged-only (date format applied)
                            alt_name = fname.rsplit(".", 1)[0] + "_merged_only.xlsx"
                            bout = io.BytesIO()
                            with pd.ExcelWriter(
                                bout,
                                engine="xlsxwriter",
                                datetime_format="dd-mmm-yy",
                                date_format="dd-mmm-yy",
                            ) as writer:
                                merged_df.to_excel(writer, index=False, sheet_name="Merged")
                            bout.seek(0)
                            zf.writestr(alt_name, bout.read())

                zipbuf.seek(0)
                st.download_button(
                    label="📦 Download ZIP (each original + merged sheet)",
                    data=zipbuf.getvalue(),
                    file_name="merged_&_original.zip",
                    mime="application/zip",
                )

                if any(f.name.lower().endswith(".xls") for f, _ in file_objs):
                    st.info("For .xls files, original formatting cannot be preserved; those are provided as new .xlsx files with the merged sheet.")

st.markdown("---")

# === Section 1: Data Cleaning ===
st.header("Data Cleaning")
clean_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="clean_uploader")

if clean_file is not None:
    try:
        clean_xls = pd.ExcelFile(clean_file)
        clean_sheets = clean_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file.")
        st.exception(e)
        clean_xls = None
        clean_sheets = []

    if clean_sheets:
        selected = st.multiselect("Select sheet(s) to process", clean_sheets, key="clean_select_ms")
        st.write("")
        run_clean = st.button("Run", key="run_clean")

        if run_clean:
            if not selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                cleaned_by_sheet = {}
                for sheet in selected:
                    raw_df = clean_xls.parse(sheet_name=sheet)
                    cleaned = clean_malaria_data(raw_df)
                    # preview copy without time
                    cleaned_display = put_comment_first(_df_display_without_time(cleaned.copy()))
                    # file copy with COMMENT moved last (data preserved)
                    cleaned_file = put_comment_last(cleaned.copy())

                    cleaned_by_sheet[sheet] = {
                        "display": cleaned_display,
                        "file": cleaned_file,
                    }

                st.subheader("⚠️ Error Rows Preview (by sheet)")
                tabs = st.tabs(selected)

                def _build_error_summary_by_column(error_series: pd.Series) -> pd.DataFrame:
                    counts = {}
                    for comment in error_series.astype(str):
                        parts = [p.strip() for p in comment.split(";") if p.strip()]
                        for p in parts:
                            if "[" in p and "]" in p:
                                col = p.split("[", 1)[1].split("]", 1)[0]
                                counts[col] = counts.get(col, 0) + 1
                    if not counts:
                        return pd.DataFrame(columns=["Column", "Error Count"])
                    return (
                        pd.DataFrame(list(counts.items()), columns=["Column", "Error Count"])
                        .sort_values("Error Count", ascending=False)
                        .reset_index(drop=True)
                    )

                for tab, sheet in zip(tabs, selected):
                    with tab:
                        display_df = cleaned_by_sheet[sheet]["display"]
                        error_df = display_df[display_df.get("COMMENT", "").astype(str).str.strip() != ""]
                        if error_df.empty:
                            st.success(f"✅ {sheet}: No data quality issues found.")
                        else:
                            st.dataframe(error_df.head(50), use_container_width=True)
                            st.info(f"{sheet} — Total error rows: {len(error_df)}")
                            summary_df = _build_error_summary_by_column(error_df["COMMENT"])
                            if not summary_df.empty:
                                st.markdown("**📊 Error Summary by Column**")
                                st.dataframe(summary_df, use_container_width=True)

                # Build output by replacing only processed sheets (preserving untouched formatting)
                st.subheader("⬇️ Download")
                try:
                    original_bytes = clean_file.getvalue() if hasattr(clean_file, "getvalue") else clean_file.read()
                    wb = load_workbook(io.BytesIO(original_bytes))

                    for sheet in clean_xls.sheet_names:
                        if sheet in cleaned_by_sheet:
                            try:
                                idx = wb.sheetnames.index(sheet)
                            except ValueError:
                                idx = len(wb.sheetnames)
                            if sheet in wb.sheetnames:
                                wb.remove(wb[sheet])
                            ws = wb.create_sheet(title=sheet, index=idx)
                            for r in dataframe_to_rows(cleaned_by_sheet[sheet]["file"], index=False, header=True):
                                ws.append(r)
                            # NEW: make dates display as DD-MMM-YY
                            _format_worksheet_dates(ws, cleaned_by_sheet[sheet]["file"])

                    out = io.BytesIO()
                    wb.save(out); out.seek(0)
                    st.download_button(
                        label="📥 Download Cleaned Workbook (preserve untouched formatting)",
                        data=out.getvalue(),
                        file_name="malaria_cleaned_all.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error("❌ Failed to generate the cleaned workbook.")
                    st.exception(e)

st.markdown("---")

# === Section 2: Indicator Calculation ===
st.header("Indicator Calculation")
ind_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="ind_uploader")

if ind_file is not None:
    try:
        ind_xls = pd.ExcelFile(ind_file)
        ind_sheets = ind_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file.")
        st.exception(e)
        ind_xls = None
        ind_sheets = []

    if ind_sheets:
        ind_selected = st.multiselect("Select sheet(s) to process", ind_sheets, key="ind_select_ms")
        st.write("")
        run_ind = st.button("Run", key="run_ind")

        if run_ind:
            if not ind_selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                outputs = {}
                errors = []
                for sheet in ind_selected:
                    try:
                        raw_df = ind_xls.parse(sheet_name=sheet)
                        out_df = compute_indicators(raw_df.copy())
                        outputs[sheet] = out_df
                    except Exception as e:
                        errors.append((sheet, e))

                if outputs:
                    st.subheader("👀 Preview (first 50 rows)")
                    tabs = st.tabs(list(outputs.keys()))
                    for tab, sheet in zip(tabs, outputs.keys()):
                        with tab:
                            st.caption(f"Sheet: {sheet}")
                            # display without time
                            st.dataframe(_df_display_without_time(outputs[sheet]).head(50), use_container_width=True)

                if errors:
                    st.warning("Some sheets failed to process:")
                    for sheet, e in errors:
                        with st.expander(f"Details: {sheet}"):
                            st.exception(e)

                # Build output by replacing only processed sheets (preserving untouched formatting)
                if outputs:
                    st.subheader("⬇️ Download")
                    try:
                        original_bytes = ind_file.getvalue() if hasattr(ind_file, "getvalue") else ind_file.read()
                        wb = load_workbook(io.BytesIO(original_bytes))

                        for sheet in ind_xls.sheet_names:
                            if sheet in outputs:
                                try:
                                    idx = wb.sheetnames.index(sheet)
                                except ValueError:
                                    idx = len(wb.sheetnames)
                                if sheet in wb.sheetnames:
                                    wb.remove(wb[sheet])
                                ws = wb.create_sheet(title=sheet, index=idx)
                                for r in dataframe_to_rows(outputs[sheet], index=False, header=True):
                                    ws.append(r)
                                # NEW: format date columns for display
                                _format_worksheet_dates(ws, outputs[sheet])

                        out = io.BytesIO()
                        wb.save(out); out.seek(0)
                        st.download_button(
                            label="📥 Download Indicators Workbook (preserve untouched formatting)",
                            data=out.getvalue(),
                            file_name="malaria_indicators_all.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error("❌ Failed to generate the indicators workbook.")
                        st.exception(e)
