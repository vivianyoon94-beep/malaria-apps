import argparse
import io
import os
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd

# Optional import for preserving formatting when injecting into .xlsx
try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    _HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAVE_OPENPYXL = False


# ---------- helpers ----------
def _normalize_headers(df: pd.DataFrame) -> List[str]:
    return [str(c).strip().lower() for c in df.columns]

def _validate_headers_match(dfs: List[pd.DataFrame]) -> None:
    if not dfs:
        raise ValueError("No dataframes selected")
    base = set(_normalize_headers(dfs[0]))
    for i, d in enumerate(dfs[1:], start=2):
        if set(_normalize_headers(d)) != base:
            raise ValueError(
                f"Headers do not match across all selected sheets (mismatch near input #{i})."
            )

def _normalize_screening_date_like_cleaning(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replicates the SCREENING_DATE normalization from the cleaner:
    - Accepts only strict formats (no permissive fallback)
    - Leaves invalid/unparsable cells unchanged
    - Converts valid dates to string 'DD-MMM-YY'
    This mirrors the logic used in Malaria_Data_Cleaning.clean_malaria_data. :contentReference[oaicite:0]{index=0}
    """
    from datetime import datetime as _dt

    # case-insensitive lookup
    col_map = {c.lower(): c for c in df.columns}
    date_col = col_map.get('screening_date')
    if not date_col:
        return df

    start_date = pd.to_datetime('2024-01-01')
    today      = pd.to_datetime(pd.Timestamp.today().strftime('%Y-%m-%d'))

    strict_formats = [
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d',
        '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y', '%d-%m-%Y',
        '%m/%d/%y', '%m-%d-%y', '%d/%m/%y', '%d-%m-%y'
    ]  # same idea as the cleaner’s strict list. :contentReference[oaicite:1]{index=1}

    for idx, val in df[date_col].items():
        if pd.isna(val) or str(val).strip() == '':
            continue

        valid, parsed = False, None

        # Already a Timestamp from Excel
        if isinstance(val, pd.Timestamp):
            parsed, valid = val, True
        else:
            s = str(val).strip()
            for fmt in strict_formats:
                try:
                    parsed = _dt.strptime(s, fmt)
                    if 2024 < parsed.year < 2100:
                        valid = True
                        break
                except Exception:
                    continue

        if not valid:
            # keep original cell (merge step should not inject comments)
            continue

        # range check (same policy; keep original if outside)
        if not (start_date <= pd.to_datetime(parsed) <= today):
            continue

        # normalize to 'DD-MMM-YY'
        df.at[idx, date_col] = pd.to_datetime(parsed).strftime('%d-%b-%y')

    return df


# ---------- core ----------
def merge_across_files(file_sheet_map: Dict[str, Iterable[str]]) -> pd.DataFrame:
    """Merge rows across multiple files/sheets. Adds DATA_SOURCE and FILE_SOURCE."""
    parts: List[pd.DataFrame] = []
    for fname, sheets in file_sheet_map.items():
        xls = pd.ExcelFile(fname)
        for s in sheets:
            df = xls.parse(sheet_name=s)
            df.columns = [str(c) for c in df.columns]
            # Normalize date like in the cleaner so downstream cleaning keeps the same format
            df = _normalize_screening_date_like_cleaning(df)  # <-- new
            df["DATA_SOURCE"] = s
            df["FILE_SOURCE"] = Path(fname).name
            parts.append(df)

    if not parts:
        raise ValueError("No sheets selected.")

    _validate_headers_match(parts)

    meta_cols = {"DATA_SOURCE", "FILE_SOURCE"}
    canonical = [c for c in parts[0].columns if c not in meta_cols]

    normalized = []
    for df in parts:
        mapping = {str(c).strip().lower(): c for c in df.columns}
        ordered = [mapping[str(c).strip().lower()] for c in canonical]
        tmp = df[ordered + ["DATA_SOURCE", "FILE_SOURCE"]].copy()
        normalized.append(tmp)

    # Concatenate without coercing datetime to .dt.date (we now normalized as strings)
    merged = pd.concat(normalized, ignore_index=True)
    return merged


def write_merged_only(path: str, merged_df: pd.DataFrame, sheet_name: str = "Merged") -> None:
    # Using XlsxWriter gives nice display for any remaining real date cells
    with pd.ExcelWriter(
        path,
        engine="xlsxwriter",
        datetime_format="dd-mmm-yy",
        date_format="dd-mmm-yy",
    ) as writer:
        merged_df.to_excel(writer, index=False, sheet_name=sheet_name)


def write_zip_injected(file_sheet_map: Dict[str, Iterable[str]], merged_df: pd.DataFrame, zip_path: str) -> None:
    """Create a ZIP where each original .xlsx gets a merged sheet appended, preserving formatting.
       For .xls inputs, include a new .xlsx containing only the merged sheet.
    """
    if not _HAVE_OPENPYXL:
        raise RuntimeError(
            "openpyxl not available; cannot preserve formatting for .xlsx. Add 'openpyxl' to requirements."
        )

    from datetime import date as _dt_date

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for src_path in file_sheet_map.keys():
            fname = os.path.basename(src_path)

            if fname.lower().endswith(".xlsx"):
                with open(src_path, "rb") as f:
                    orig_bytes = f.read()

                wb = load_workbook(io.BytesIO(orig_bytes))

                # Unique sheet name and put the merged sheet FIRST
                base_name = "Merged"
                name = base_name
                counter = 1
                while name in wb.sheetnames:
                    counter += 1
                    name = f"{base_name}_{counter}"
                ws = wb.create_sheet(title=name, index=0)

                # Write header + rows
                for r in dataframe_to_rows(merged_df, index=False, header=True):
                    ws.append(r)

                # Excel display format for TRUE date/datetime cells (if any remain)
                date_cols_idx = []
                for j, col in enumerate(merged_df.columns, start=1):
                    col_series = merged_df[col]
                    if pd.api.types.is_datetime64_any_dtype(col_series) or col_series.map(
                        lambda v: isinstance(v, (_dt_date, pd.Timestamp))
                    ).any():
                        date_cols_idx.append(j)
                for j in date_cols_idx:
                    for col_cells in ws.iter_cols(min_col=j, max_col=j, min_row=2):
                        for cell in col_cells:
                            cell.number_format = "DD-MMM-YY"

                out = io.BytesIO()
                wb.save(out)
                out.seek(0)
                zf.writestr(fname, out.read())

            else:
                # .xls fallback: cannot preserve original formatting; deliver an xlsx with merged-only
                alt_name = os.path.splitext(fname)[0] + "_merged_only.xlsx"
                bout = io.BytesIO()
                with pd.ExcelWriter(
                    bout,
                    engine="xlsxwriter",
                    datetime_format="dd-mmm-yy",
                    date_format="dd-mmm-yy",
                ) as writer:
                    merged_df.to_excel(writer, index=False, sheet_name="Merged")
                bout.seek(0)
                zf.writestr(alt_name, bout.read())


# ---------- CLI ----------
def parse_select_args(select_args: List[str], files: List[str]) -> Dict[str, List[str]]:
    """Parse selections of the form:
       --select "file1.xlsx:SheetA,SheetB"  --select "file2.xlsx:Data"
       If no --select provided, default to all sheets for each file.
    """
    mapping: Dict[str, List[str]] = {}
    if not select_args:
        for f in files:
            mapping[f] = pd.ExcelFile(f).sheet_names
        return mapping

    wanted: Dict[str, List[str]] = {}
    for item in select_args:
        item = item.strip()
        if not item or ":" not in item:
            raise ValueError(f"Bad --select value: {item!r}. Use 'file.xlsx:Sheet1,Sheet2'")
        fname, sheets_str = item.split(":", 1)
        fname = fname.strip()
        if fname not in files:
            raise ValueError(f"--select references unknown file: {fname}")
        sheet_list = [s.strip() for s in sheets_str.split(",") if s.strip()]
        if not sheet_list:
            raise ValueError(f"No sheets provided in --select for {fname}")
        wanted[fname] = sheet_list

    for f in files:
        mapping[f] = wanted.get(f, pd.ExcelFile(f).sheet_names)
    return mapping


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Merge sheets across one or more Excel files.")
    ap.add_argument("files", nargs="+", help="Paths to input Excel files (.xlsx or .xls)")
    ap.add_argument("-o", "--output", default="merged.xlsx", help="Path to merged-only workbook to write")
    ap.add_argument("--zip", dest="zip_path", help="Create a ZIP that injects the merged sheet back into each original file")
    ap.add_argument("--select", action="append", help="Select sheets per file: 'file.xlsx:Sheet1,Sheet2' (repeatable)")
    args = ap.parse_args(argv)

    file_sheet_map = parse_select_args(args.select or [], args.files)
    merged_df = merge_across_files(file_sheet_map)

    write_merged_only(args.output, merged_df)
    print(f"✅ Wrote merged-only workbook: {args.output}")

    if args.zip_path:
        write_zip_injected(file_sheet_map, merged_df, args.zip_path)
        print(f"📦 Wrote ZIP bundle with injected merged sheets: {args.zip_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
