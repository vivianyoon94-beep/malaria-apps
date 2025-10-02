
import io
import zipfile
import pandas as pd
import streamlit as st

# Import library functions for other sections
from Malaria_Data_Cleaning import clean_malaria_data
from Malaria_Indicator import compute_indicators

# For preserving untouched formatting when injecting merged sheet
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- Utilities ----------------
def put_comment_first(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[["COMMENT"] + [c for c in df.columns if c != "COMMENT"]]

def put_comment_last(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[[c for c in df.columns if c != "COMMENT"] + ["COMMENT"]]

def sanitize_sheet_name(name: str) -> str:
    bad = '[]:*?/\\'
    table = {ord(ch): ' ' for ch in bad}
    name = str(name).translate(table).strip() or "Sheet"
    return name[:31]

def _normalize_headers(df: pd.DataFrame):
    return [str(c).strip().lower() for c in df.columns]

def _validate_headers_match(dfs):
    if not dfs:
        return False, "No dataframes selected"
    base = set(_normalize_headers(dfs[0]))
    for i, d in enumerate(dfs[1:], start=2):
        if set(_normalize_headers(d)) != base:
            return False, i
    return True, None

def _strip_time_from_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Only convert proper datetime64[ns] columns to date
    for col, dtype in df.dtypes.items():
        if str(dtype).startswith("datetime64"):
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
    return df

# Optional helper: normalize SCREENING_DATE as 'DD-MMM-YY' for display
import re as _re
from datetime import datetime as _dt
_ISO_DATE = _re.compile(r'^\d{4}[-/]\d{2}[-/]\d{2}(?: \d{2}:\d{2}:\d{2})?$')
def force_screening_date_strings(df: pd.DataFrame) -> pd.DataFrame:
    date_col = next((c for c in df.columns if c.strip().lower() == "screening_date"), None)
    if not date_col:
        return df
    def _fmt(v):
        if isinstance(v, (pd.Timestamp, _dt)):
            return pd.to_datetime(v).strftime('%d-%b-%y')
        if isinstance(v, str) and _ISO_DATE.match(v.strip()):
            dt = pd.to_datetime(v, errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%d-%b-%y')
        return v
    df[date_col] = df[date_col].map(_fmt)
    return df

# ----------------- App -----------------
st.set_page_config(page_title="🦟 Malaria Apps", layout="wide")
st.title("🦟 Malaria Apps")

# --- Section 0: Sheet Merger (single or multiple files) ---
st.header("Sheet Merger (single or multiple files)")

merge_files = st.file_uploader(
    "Upload one or more Excel files to merge",
    type=["xlsx", "xls"],
    key="merge_uploader_multi",
    accept_multiple_files=True,
)

if merge_files:
    # Build per-file ExcelFile objects and sheet lists
    file_objs = []
    for f in merge_files:
        try:
            xls = pd.ExcelFile(f)
            file_objs.append((f, xls))
        except Exception as e:
            st.error(f"❌ Could not read: {f.name}")
            st.exception(e)
            file_objs.append((f, None))

    # Per-file sheet selection UI
    selections = {}
    for f, xls in file_objs:
        st.subheader(f"File: {f.name}")
        if xls is None:
            continue
        sheets = xls.sheet_names
        chosen = st.multiselect(f"Select sheet(s) from **{f.name}** to include", sheets, key=f"sel_{f.name}")
        selections[f.name] = {
            "sheets": chosen,
            "xls": xls,
            "file": f,
        }

    st.write("")
    run_merge = st.button("Run", key="run_merge_multi")

    if run_merge:
        # Gather selected DataFrames across all files
        parts = []
        for fname, meta in selections.items():
            xls = meta.get("xls")
            chosen = meta.get("sheets", [])
            if xls is None or not chosen:
                continue
            for s in chosen:
                df = xls.parse(sheet_name=s)
                df.columns = [str(c) for c in df.columns]
                df["DATA_SOURCE"] = s
                df["FILE_SOURCE"] = fname
                parts.append(df)

        if not parts:
            st.warning("Please select at least **one sheet** to merge across the uploaded files.")
        else:
            # Validate headers across all chosen parts
            ok, info = _validate_headers_match(parts)
            if not ok:
                if isinstance(info, int):
                    st.error(f"Headers do not match across all selected sheets. Mismatch around input #{info}.")
                else:
                    st.error(str(info))
            else:
                # Canonical order = first part's original columns (excluding the two metadata columns)
                meta_cols = {"DATA_SOURCE", "FILE_SOURCE"}
                canonical = [c for c in parts[0].columns if c not in meta_cols]
                # Reorder each part to match canonical
                normalized = []
                for df in parts:
                    mapping = {str(c).strip().lower(): c for c in df.columns}
                    ordered = [mapping[str(c).strip().lower()] for c in canonical]
                    tmp = df[ordered + ["DATA_SOURCE", "FILE_SOURCE"]].copy()
                    normalized.append(tmp)

                merged_df = pd.concat(normalized, ignore_index=True)
                merged_df = _strip_time_from_datetime_columns(merged_df)

                st.subheader("👀 Merged preview (first 50 rows)")
                st.dataframe(merged_df.head(50), use_container_width=True)

                # Download options
                st.subheader("⬇️ Download")
                try:
                    # 1) Merged-only workbook
                    out = io.BytesIO()
                    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
                        merged_df.to_excel(writer, index=False, sheet_name="Merged")
                    out.seek(0)
                    st.download_button(
                        label="📥 Download merged-only workbook",
                        data=out.getvalue(),
                        file_name="merged_across_files.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    # 2) ZIP: inject merged into each original file (xlsx keeps formatting)
                    zipbuf = io.BytesIO()
                    with zipfile.ZipFile(zipbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for f, xls in file_objs:
                            fname = f.name
                            if fname.lower().endswith(".xlsx"):
                                # Preserve formatting using openpyxl
                                orig_bytes = f.getvalue() if hasattr(f, "getvalue") else f.read()
                                wb = load_workbook(io.BytesIO(orig_bytes))
                                # unique merged sheet name per file
                                base_name = "Merged"
                                name = base_name
                                counter = 1
                                while name in wb.sheetnames:
                                    counter += 1
                                    name = f"{base_name}_{counter}"
                                ws = wb.create_sheet(title=name)
                                for r in dataframe_to_rows(merged_df, index=False, header=True):
                                    ws.append(r)
                                fout = io.BytesIO()
                                wb.save(fout); fout.seek(0)
                                zf.writestr(fname, fout.read())
                            else:
                                # .xls: cannot preserve formatting; provide a new .xlsx with merged-only
                                alt_name = fname.rsplit(".", 1)[0] + "_merged_only.xlsx"
                                bout = io.BytesIO()
                                with pd.ExcelWriter(bout, engine="xlsxwriter") as writer:
                                    merged_df.to_excel(writer, index=False, sheet_name="Merged")
                                bout.seek(0)
                                zf.writestr(alt_name, bout.read())

                    zipbuf.seek(0)
                    st.download_button(
                        label="📦 Download ZIP (each original + merged sheet)",
                        data=zipbuf.getvalue(),
                        file_name="merged_injected_bundle.zip",
                        mime="application/zip",
                    )

                    # Note for users with .xls
                    if any(f.name.lower().endswith(".xls") for f, _ in file_objs):
                        st.info("For .xls files, formatting cannot be preserved. Those entries are provided as new .xlsx files containing only the merged sheet.")
                except Exception as e:
                    st.error("❌ Failed to build the downloads.")
                    st.exception(e)

st.markdown("---")

# --- Section 1: Data Cleaning ---
st.header("Data Cleaning")

clean_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="clean_uploader")

if clean_file is not None:
    try:
        clean_xls = pd.ExcelFile(clean_file)
        clean_sheets = clean_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file. Please check the format.")
        st.exception(e)
        clean_xls = None
        clean_sheets = []

    if clean_sheets:
        clean_selected = [*st.session_state.get("clean_selected", [])]  # backward compat (not used)
        clean_selected = st.multiselect("Select sheet(s) to process", clean_sheets, default=clean_selected, key="clean_select_ms")
        st.write("")  # small gap
        run_clean = st.button("Run", key="run_clean")

        if run_clean:
            if not clean_selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                # Process selected sheets to generate cleaned DataFrames
                cleaned_by_sheet = {}
                for sheet in clean_selected:
                    raw_df  = clean_xls.parse(sheet_name=sheet)
                    cleaned = clean_malaria_data(raw_df)
                    cleaned = force_screening_date_strings(cleaned)
                    cleaned_by_sheet[sheet] = {
                        "display": put_comment_first(cleaned.copy()),
                        "file":    put_comment_last(cleaned.copy()),
                    }

                st.subheader("⚠️ Error Rows Preview (by sheet)")
                tabs = st.tabs(clean_selected)

                def _build_error_summary_by_column(error_series: pd.Series) -> pd.DataFrame:
                    counts = {}
                    for comment in error_series.astype(str):
                        parts = [p.strip() for p in comment.split(';') if p.strip()]
                        for p in parts:
                            if '[' in p and ']' in p:
                                col = p.split('[', 1)[1].split(']', 1)[0]
                                counts[col] = counts.get(col, 0) + 1
                    if not counts:
                        return pd.DataFrame(columns=['Column', 'Error Count'])
                    return (
                        pd.DataFrame(list(counts.items()), columns=['Column', 'Error Count'])
                        .sort_values('Error Count', ascending=False)
                        .reset_index(drop=True)
                    )

                for tab, sheet in zip(tabs, clean_selected):
                    with tab:
                        display_df = cleaned_by_sheet[sheet]["display"]
                        error_df = display_df[display_df.get("COMMENT", "").astype(str).str.strip() != ""]
                        if error_df.empty:
                            st.success(f"✅ {sheet}: No data quality issues found.")
                        else:
                            st.dataframe(error_df.head(50), use_container_width=True)
                            st.info(f"{sheet} — Total error rows: {len(error_df)}")
                            summary_df = _build_error_summary_by_column(error_df["COMMENT"])
                            if not summary_df.empty:
                                st.markdown("**📊 Error Summary by Column**")
                                st.dataframe(summary_df, use_container_width=True)

                # Download workbook: include ALL sheets (selected cleaned, unselected untouched)
                st.subheader("⬇️ Download")
                try:
                    original_bytes = clean_file.getvalue() if hasattr(clean_file, "getvalue") else clean_file.read()
                    wb = load_workbook(io.BytesIO(original_bytes))

                    for sheet in clean_xls.sheet_names:
                        if sheet in cleaned_by_sheet:
                            try:
                                idx = wb.sheetnames.index(sheet)
                            except ValueError:
                                idx = len(wb.sheetnames)
                            if sheet in wb.sheetnames:
                                wb.remove(wb[sheet])
                            ws = wb.create_sheet(title=sheet, index=idx)
                            for r in dataframe_to_rows(cleaned_by_sheet[sheet]["file"], index=False, header=True):
                                ws.append(r)

                    out = io.BytesIO()
                    wb.save(out)
                    out.seek(0)

                    st.download_button(
                        label=f"📥 Download Cleaned Workbook (preserve untouched formatting)",
                        data=out.getvalue(),
                        file_name="malaria_cleaned_all.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error("❌ Failed to generate the cleaned workbook.")
                    st.exception(e)

st.markdown("---")

# --- Section 2: Indicator Calculation ---
st.header("Indicator Calculation")

ind_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="ind_uploader")

if ind_file is not None:
    try:
        ind_xls = pd.ExcelFile(ind_file)
        ind_sheets = ind_xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file. Please check the format.")
        st.exception(e)
        ind_xls = None
        ind_sheets = []

    if ind_sheets:
        ind_selected = st.multiselect("Select sheet(s) to process", ind_sheets, key="ind_select_ms")
        st.write("")  # small gap
        run_ind = st.button("Run", key="run_ind")

        if run_ind:
            if not ind_selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                outputs = {}
                errors = []
                for sheet in ind_selected:
                    try:
                        raw_df = ind_xls.parse(sheet_name=sheet)
                        out_df = compute_indicators(raw_df.copy())
                        outputs[sheet] = out_df
                    except Exception as e:
                        errors.append((sheet, e))

                if outputs:
                    st.subheader("👀 Preview (first 50 rows)")
                    tabs = st.tabs(list(outputs.keys()))
                    for tab, sheet in zip(tabs, outputs.keys()):
                        with tab:
                            st.caption(f"Sheet: {sheet}")
                            st.dataframe(outputs[sheet].head(50), use_container_width=True)

                if errors:
                    st.warning("Some sheets failed to process:")
                    for sheet, e in errors:
                        with st.expander(f"Details: {sheet}"):
                            st.exception(e)

                # --- Build downloadable workbook, preserving UNTOUCHED sheets via openpyxl ---
                if outputs or ind_selected:
                    st.subheader("⬇️ Download")
                    try:
                        original_bytes = ind_file.getvalue() if hasattr(ind_file, "getvalue") else ind_file.read()
                        wb = load_workbook(io.BytesIO(original_bytes))

                        for sheet in ind_xls.sheet_names:
                            if sheet in outputs:
                                try:
                                    idx = wb.sheetnames.index(sheet)
                                except ValueError:
                                    idx = len(wb.sheetnames)
                                if sheet in wb.sheetnames:
                                    wb.remove(wb[sheet])
                                ws = wb.create_sheet(title=sheet, index=idx)

                                for r in dataframe_to_rows(outputs[sheet], index=False, header=True):
                                    ws.append(r)

                        out = io.BytesIO()
                        wb.save(out)
                        out.seek(0)

                        st.download_button(
                            label=f"📥 Download Indicators Workbook (preserve untouched formatting)",
                            data=out.getvalue(),
                            file_name="malaria_indicators_all.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error("❌ Failed to generate the indicators workbook.")
                        st.exception(e)
